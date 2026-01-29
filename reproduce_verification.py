import pandas as pd
import openpyxl
from pathlib import Path
import os
import sys

# Add root to path
sys.path.append(os.getcwd())

from sheet_verifier.main import main

def create_mock_data():
    base_dir = Path("test_verification_env")
    base_dir.mkdir(exist_ok=True)
    
    # 1. Master List
    # Note: 'id' key here must match column mapping logic in master_loader (inv id, invoice no etc)
    master_data = [
        {'Invoice No': 'INV-100', 'Amount': 5000, 'Quantity': 100, 'Pallets': 10, 'ID': 'INV-100'},
        {'Invoice No': 'INV-101', 'Amount': 200, 'Quantity': 20, 'Pallets': 2, 'ID': 'INV-101'},
    ]
    master_path = base_dir / "MasterList.csv"
    pd.DataFrame(master_data).to_csv(master_path, index=False)
    
    # 2. Valid Invoice (INV-100)
    wb = openpyxl.Workbook()
    
    # Invoice Sheet
    ws1 = wb.active
    ws1.title = "Invoice"
    ws1['A1'] = "Invoice No"
    # ws1['B1'] = "INV-100" # Old placement (Right)
    ws1['A2'] = "INV-100"   # New placement (Below) - Testing enhanced search
    
    ws1['A10'] = "Total :"
    ws1['B10'] = 5000   # Amount
    ws1['C10'] = 100    # Quantity
    ws1['D10'] = "10 Pallets"
    
    # ws1['B1'] = "Amount" # Header helper removed for ID check clarity, relies on cell proximity
    # Re-add format headers logic check if needed? 
    # Extractor uses proximity to "Amount"/"Quantity" strings.
    ws1['B9'] = "Amount"
    ws1['C9'] = "Quantity"
    
    # Packing List Sheet
    ws2 = wb.create_sheet("Packing list")
    ws2['A5'] = "Total of :"
    ws2['B5'] = 100 # Quantity
    ws2['C5'] = "Total 10 Pallets"
    
    ws2['B1'] = "Quantity"
    
    wb.save(base_dir / "INV-100_Valid.xlsx")
    
    # 3. Invalid Invoice (INV-101) - Mismatch in Packing List
    wb2 = openpyxl.Workbook()
    ws3 = wb2.active
    ws3.title = "Invoice"
    ws3['A10'] = "Total:"
    ws3['B10'] = 200
    ws3['C10'] = 20
    ws3['D10'] = "2 Pallets"
    ws3['B1'] = "Amount"
    ws3['C1'] = "Quantity"
    
    ws4 = wb2.create_sheet("Packing list")
    ws4['A5'] = "Total:"
    ws4['B5'] = 15 # WARN: Quantity Mismatch (Exp 20)
    ws4['C5'] = "Total 2 Pallets"
    ws4['B1'] = "Quantity"
    
    wb2.save(base_dir / "INV-101_Invalid.xlsx")
    
    return base_dir, master_path

def test_run():
    folder, master = create_mock_data()
    print(f"Created mock env at {folder}")
    
    # Run Main
    import sys
    # Mock sys.argv
    sys.argv = ["main.py", "--folder", str(folder), "--master", str(master)]
    
    try:
        main()
    except SystemExit:
        pass
        
    # Check Result
    report_path = folder / "verification_report.csv"
    if report_path.exists():
        print("\n--- Report Content ---")
        with open(report_path, 'r') as f:
            print(f.read())
    else:
        print("Report not found!")

if __name__ == "__main__":
    test_run()
