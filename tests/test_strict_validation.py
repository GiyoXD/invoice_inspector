
import sys
import os
import pandas as pd
import openpyxl
from pathlib import Path

# Add project root to sys.path
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

from sheet_verifier.verifier import InvoiceVerifier
from sheet_verifier.master_loader import load_master_list
from core.exceptions import ErrorCode

def create_dummy_master_list(path):
    data = {
        'Invoice No': ['INV-001'],
        'Amount': [1000.0],
        'Quantity': [500.0],
        'Pallets': [10.0],
        'PCS': [5000],
        'Net Weight': [200.0],
        'Gross Weight': [220.0],
        'CBM': [1.5]
    }
    df = pd.DataFrame(data)
    df.to_csv(path, index=False)
    print(f"Created dummy Master List at {path}")

def create_test_excel(path, scenario):
    wb = openpyxl.Workbook()
    
    # 1. Invoice Sheet (Always present, but maybe missing cols)
    ws_inv = wb.active
    ws_inv.title = "Invoice"
    
    # Determine columns and total row values
    if scenario == 'valid':
        headers = ["Invoice No", "Amount", "Quantity", "Pallet No"]
        total_vals = ["Total:", 1000, 500, 10]
    elif scenario == 'missing_col':
        headers = ["Invoice No", "Quantity", "Pallet No"] # Missing Amount
        total_vals = ["Total:", 500, 10]
    elif scenario == 'invalid_value':
        headers = ["Invoice No", "Amount", "Quantity", "Pallet No"]
        total_vals = ["Total:", "TBD", 500, 10] # Invalid Amount "TBD"
    elif scenario == 'partition_mismatch':
        headers = ["Invoice No", "Amount", "Quantity", "Pallet No"]
        total_vals = ["Total:", 900, 500, 10] # Amount mismatch (1000 expected)
    
    ws_inv.append(headers)
    ws_inv.append(["INV-001", "", "", ""]) # Dummy Line Item
    ws_inv.append(total_vals) # Total Row with Data
    
    # 2. Contract Sheet (Required for Amount/Qty)
    ws_con = wb.create_sheet("Contract")
    ws_con.append(["Invoice No", "Amount", "Quantity"])
    ws_con.append(["INV-001", "", ""])
    if scenario == 'partition_mismatch':
        ws_con.append(["Total:", 1000, 500]) # Valid here, but Invoice has 900 -> Mismatch
    else:
        ws_con.append(["Total:", 1000, 500])

    # 3. Packing List Sheet (Required for Qty/Pallet/Pcs/Net/Gross/CBM)
    ws_pak = wb.create_sheet("Packing List")
    ws_pak.append(["Invoice No", "Quantity", "Pallet", "PCS", "Net Weight", "Gross Weight", "CBM"])
    ws_pak.append(["INV-001", "", "", "", "", "", ""])
    ws_pak.append(["Total:", 500, 10, 5000, 200, 220, 1.5]) # Data in Total Row

    wb.save(path)
    print(f"Created test Excel ({scenario}) at {path}")

def run_test():
    base_dir = Path("tests/temp_strict")
    base_dir.mkdir(parents=True, exist_ok=True)
    
    master_path = base_dir / "MasterList.csv"
    create_dummy_master_list(master_path)
    
    master_data = load_master_list(master_path)
    verifier = InvoiceVerifier()
    
    scenarios = ['valid', 'missing_col', 'invalid_value', 'partition_mismatch']
    
    for sc in scenarios:
        print(f"\n--- Testing Scenario: {sc} ---")
        xlsx_path = base_dir / f"test_{sc}.xlsx"
        create_test_excel(xlsx_path, sc)
        
        result = verifier.verify_file(xlsx_path, master_data.get('INV-001'))
        
        print(f"Status: {result['status']}")
        for d in result['details']:
            print(f" - {d}")

if __name__ == "__main__":
    run_test()
