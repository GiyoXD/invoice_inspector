
import re
from pathlib import Path
from openpyxl import load_workbook
from typing import List, Dict, Set, Optional
from core.config import load_mapping_config
from core.models import ExtractedInvoice, VerificationStatus
from core.regex_utils import regex_extract_number, regex_extract

# Config
BLACKLIST_TERMS = ["buffalo", "cow", "leather"]

def find_invoice_sheet(wb):
    """Finds a sheet named like 'invoice', 'inv', etc."""
    sheet_names = wb.sheetnames
    for name in sheet_names:
        lower_name = name.lower()
        if 'invoice' in lower_name or 'inv' in lower_name:
            return wb[name]
    return None

def find_contract_sheet(wb):
    """Finds a sheet named like 'contract' or 'ct'."""
    for sheet in wb:
        title = sheet.title.lower().strip()
        if 'contract' in title: return sheet
        if title == 'ct': return sheet
        if title.startswith('ct ') or title.startswith('ct-') or title.startswith('ct&') or title.startswith('ct_'):
            return sheet
        if title.endswith(' ct'): return sheet
    return None

def find_all_packing_list_sheets(wb):
    """Finds ALL sheets that look like a packing list."""
    sheets = []
    for sheet in wb:
        title = sheet.title.lower()
        # Must contain 'pack' or 'packing' - 'detail' alone is too generic
        if 'pack' in title or 'packing' in title:
            sheets.append(sheet)
        elif 'weight' in title and ('gross' in title or 'net' in title):
            # Allow weight-related sheets only if they have gross/net context
            sheets.append(sheet)
    return sheets

def extract_pallet_info(cell_value):
    """Extracts text to the left of 'pallet'."""
    if not isinstance(cell_value, str):
        return None
    result = regex_extract(cell_value, r'(.*?)\s*pallet', group=1)
    return result.strip() if result else None

def detect_inspectable_columns(sheet, mapping_dict) -> set:
    """
    Finds the header row by detecting a cluster of 3+ recognized headers in the same row.
    Returns a set of col_ids that are inspectable on this sheet.
    
    Uses header_text_mappings from mapping_config.json for header recognition.
    """
    # Scan first 50 rows to find header cluster
    max_row = min(50, sheet.max_row or 50)
    max_col = min(30, sheet.max_column or 30)
    
    best_row = -1
    best_col_ids = set()
    best_match_count = 0
    
    # STEP 1: Find the widest rows (most populated cells)
    row_cell_counts = {}
    row_cells_data = {}
    
    for row in range(1, max_row + 1):
        cell_count = 0
        cells = []
        
        for col in range(1, max_col + 1):
            cell_val = sheet.cell(row=row, column=col).value
            if cell_val:
                cell_count += 1
                text = str(cell_val).strip()
                text_lower = text.lower().replace('\n', ' ').replace('  ', ' ')
                cells.append(text_lower[:30])
        
        if cell_count > 0:
            row_cell_counts[row] = cell_count
            row_cells_data[row] = cells
    
    # Sort rows by cell count (widest first)
    sorted_rows = sorted(row_cell_counts.items(), key=lambda x: -x[1])
    
    # STEP 2: Check the widest rows against mapping (try top 5 widest)
    best_row = -1
    best_col_ids = set()
    best_match_count = 0
    debug_rows = {}
    
    for row, cell_count in sorted_rows[:10]:  # Check top 10 widest rows
        row_col_ids = set()
        row_matches = []
        match_count = 0
        
        for col in range(1, max_col + 1):
            cell_val = sheet.cell(row=row, column=col).value
            if not cell_val:
                continue
            
            text = str(cell_val).strip()
            text_lower = text.lower().replace('\n', ' ').replace('  ', ' ')
            
            if text_lower in mapping_dict:
                col_id = mapping_dict[text_lower]
                row_col_ids.add(col_id)
                row_matches.append(text_lower[:20])
                match_count += 1
        
        debug_rows[row] = {
            'cells': row_cells_data[row],
            'matches': row_matches,
            'count': match_count,
            'cell_count': cell_count
        }
        
        # Pick this row if it has 3+ matches and is better than current best
        if match_count >= 3 and match_count > best_match_count:
            best_row = row
            best_col_ids = row_col_ids
            best_match_count = match_count
    
    # STEP 3: Also check subheader row (row+1) for additional columns like PCS
    if best_row != -1 and best_row + 1 <= sheet.max_row:
        subheader_row = best_row + 1
        subheader_matches = []
        for col in range(1, max_col + 1):
            cell_val = sheet.cell(row=subheader_row, column=col).value
            if not cell_val:
                continue
            text = str(cell_val).strip()
            text_lower = text.lower().replace('\n', ' ').replace('  ', ' ')
            if text_lower in mapping_dict:
                col_id = mapping_dict[text_lower]
                best_col_ids.add(col_id)
                subheader_matches.append(f"{text_lower}={col_id}")
        if subheader_matches:
            print(f"    Subheader row {subheader_row}: {subheader_matches}")
    
    # Filter to only verification-relevant col_ids
    verification_cols = {'col_qty_sf', 'col_amount', 'col_pallet_count', 
                        'col_qty_pcs', 'col_net', 'col_gross', 'col_cbm'}
    inspectable = best_col_ids & verification_cols
    
    # Build detection info for Details display
    detection_info = {
        'header_row': best_row,
        'detected_cols': list(inspectable),
        'status': 'ok' if best_row != -1 else 'failed'
    }
    
    # STRICT: Log warning if no header cluster found
    if best_row == -1:
        print(f"  WARNING: No header row found in sheet '{sheet.title}' (need 3+ matching headers)")
        # DEBUG: Show widest rows and their matches
        print(f"    DEBUG: Top widest rows checked (by cell count):")
        for row, cell_count in sorted_rows[:5]:
            d = debug_rows.get(row, {})
            matches = d.get('matches', [])
            cells = d.get('cells', [])
            print(f"      Row {row}: {cell_count} cells, {len(matches)} matches -> {matches}")
            print(f"        Cells: {cells[:6]}...")
        detection_info['warning'] = f"No header row found (need 3+ matches)"
    else:
        print(f"  Detected headers (row {best_row}): {list(inspectable)}")
    
    return inspectable, detection_info

def identify_column_type(sheet, row_idx, col_idx, mapping_dict):
    """
    Looks upwards from (row_idx, col_idx) to find a header matching the mapping config.
    Returns the mapped col_id (e.g., 'col_qty_sf', 'col_amount') if found, else None.
    """
    for r in range(row_idx - 1, 0, -1):
        cell_val = sheet.cell(row=r, column=col_idx).value
        if not cell_val:
            continue
        
        text = str(cell_val).lower().strip().replace('\n', ' ')
        if text in mapping_dict:
            return mapping_dict[text]
            
        if 'total' in text and 'value' in text: return 'col_amount'
        if 'amount' in text: return 'col_amount'
             
    return None

def find_smart_total_row(sheet_values, sheet_formulas) -> int:
    """
    Identifies the best 'Total' row index.
    """
    best_row_idx = -1
    max_score = 0
    
    # Iterate rows
    # Note: sheet_values and sheet_formulas must be openpyxl sheet objects/iterators
    # To avoid iterating entire sheet, we assume reasonable size or just first 100?
    # Original logic iterated all.
    
    for row in sheet_values.iter_rows():
        row_idx = row[0].row
        has_total = False
        has_blacklist = False
        
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                v = cell.value.lower()
                for term in BLACKLIST_TERMS:
                    if term in v:
                        has_blacklist = True
                        break
                if has_blacklist: break
                
                if 'total' in v:
                    has_total = True
        
        if has_blacklist or not has_total:
            continue
            
        # Check formulas
        formula_score = 0
        try:
            f_row = sheet_formulas[row_idx]
            for cell in f_row:
                if cell.value and isinstance(cell.value, str):
                    fv = cell.value.upper()
                    if str(fv).startswith('=SUM'):
                        formula_score += 2
                    elif '+' in str(fv) and '=' in str(fv):
                         formula_score += 1
        except Exception:
            pass
            
        current_score = 1 + formula_score
        if current_score > max_score:
            max_score = current_score
            best_row_idx = row_idx
            
    return best_row_idx

def extract_packing_list_data(sheet_values, sheet_formulas, mapping_dict) -> dict:
    """Extracts data from Packing List sheet."""
    data = {
        'col_qty_pcs': 0, 'col_net': 0.0, 'col_gross': 0.0, 'col_cbm': 0.0, 'col_pallet_count': 0,
        'col_qty_sf': 0.0, 'col_amount': 0.0  # Added for per-sheet verification
    }
    row_idx = find_smart_total_row(sheet_values, sheet_formulas)
    if row_idx == -1: return {}
    
    try:
        row = sheet_values[row_idx]
        for cell in row:
             if cell.value is None: continue 
             c_type = identify_column_type(sheet_values, row_idx, cell.column, mapping_dict)
             try:
                 val = cell.value
                 if val is None: continue
                 val_str = str(val).strip()
                 num = regex_extract_number(val_str, default=0.0)

                 # Pallet Text Search
                 if 'pallet' in val_str.lower():
                     m_pal = re.search(r'(\d+)\s*[-_]?\s*pallet', val_str, re.IGNORECASE) or re.search(r'pallet\w*\s*[:\-]?\s*(\d+)', val_str, re.IGNORECASE) or re.search(r'(\d+)', val_str)
                     if m_pal:
                         p_val = int(float(m_pal.group(1)))
                         if p_val > 0: data['col_pallet_count'] = p_val
                 
                 if num == 0.0: continue

                 if c_type == 'col_qty_pcs': data['col_qty_pcs'] = int(num)
                 elif c_type == 'col_net': data['col_net'] = num
                 elif c_type == 'col_gross': data['col_gross'] = num
                 elif c_type == 'col_cbm': data['col_cbm'] = num
                 elif c_type == 'col_pallet_count': data['col_pallet_count'] = int(num)
                 elif c_type == 'col_qty_sf': data['col_qty_sf'] = num
                 elif c_type == 'col_amount': data['col_amount'] = num
             except: pass
    except Exception as e:
        print(f"Error extracting packing list row: {e}")

    # Clean result
    res = {}
    if data['col_qty_pcs'] > 0: res['col_qty_pcs'] = data['col_qty_pcs']
    if data['col_net'] > 0: res['col_net'] = round(data['col_net'], 3)
    if data['col_gross'] > 0: res['col_gross'] = round(data['col_gross'], 3)
    if data['col_cbm'] > 0: res['col_cbm'] = round(data['col_cbm'], 4)
    if data['col_pallet_count'] > 0: res['col_pallet_count'] = data['col_pallet_count']
    if data['col_qty_sf'] > 0: res['col_qty_sf'] = round(data['col_qty_sf'], 2)
    if data['col_amount'] > 0: res['col_amount'] = round(data['col_amount'], 2)
    return res

def extract_contract_data(sheet_values, sheet_formulas, mapping_dict) -> dict:
    """Extracts from Contract sheet using dynamic column detection."""
    data = {'col_qty_sf': 0.0, 'col_amount': 0.0}
    total_row_idx = find_smart_total_row(sheet_values, sheet_formulas)
    if total_row_idx == -1: 
        print("  WARNING: No Total row found in Contract sheet")
        return data

    try:
        row = sheet_values[total_row_idx]
        
        def get_float(cell):
            if cell.value:
                if isinstance(cell.value, str):
                    return regex_extract_number(cell.value, default=0.0)
                if isinstance(cell.value, (int, float)): 
                    return float(cell.value)
            return 0.0
        
        # Use dynamic column detection instead of hardcoded indices
        for cell in row:
            if cell.value is None:
                continue
            col_type = identify_column_type(sheet_values, total_row_idx, cell.column, mapping_dict)
            val = get_float(cell)
            if val > 0:
                if col_type == 'col_qty_sf':
                    data['col_qty_sf'] = val
                elif col_type == 'col_amount':
                    data['col_amount'] = val
                    
    except Exception as e:
        print(f"Error reading contract row: {e}")
    return data

def excel_data_extractor(file_path: Path) -> ExtractedInvoice:
    """Main extraction logic for a single file."""
    mapping_dict = load_mapping_config()
    
    # Initialize Model
    result = ExtractedInvoice(file_path=str(file_path), file_name=file_path.name)
    
    try:
        wb = load_workbook(file_path, data_only=True)
        wb_formulas = load_workbook(file_path, data_only=False)
    except Exception as e:
        print(f"Error loading workbook {file_path.name}: {e}")
        result.status = VerificationStatus.UNKNOWN
        return result

    total_sheet_count = len(wb.sheetnames)
    matched_sheet_count = 0

    # 1. Invoice
    inv_sheet = find_invoice_sheet(wb)
    if inv_sheet:
        matched_sheet_count += 1
        result.sheet_status['Invoice'] = True
        inv_sheet_formulas = wb_formulas[inv_sheet.title]
        
        # Detect inspectable columns for this sheet
        inv_inspectable, inv_detection = detect_inspectable_columns(inv_sheet, mapping_dict)
        result.sheets['Invoice']['target_inspect_col'] = list(inv_inspectable)
        result.sheets['Invoice']['detection_info'] = inv_detection
        result.sheets['Invoice']['sheet_title'] = inv_sheet.title
        
        data_row_idx = find_smart_total_row(inv_sheet, inv_sheet_formulas)
        if data_row_idx != -1:
             found_cols = {}
             row_cells = inv_sheet[data_row_idx]
             for cell in row_cells:
                  if cell.value is not None:
                      c_type = identify_column_type(inv_sheet, cell.row, cell.column, mapping_dict)
                      if c_type: found_cols[c_type] = cell.column
             
             def get_v(c_type):
                 col = found_cols.get(c_type)
                 if col: return inv_sheet.cell(row=data_row_idx, column=col).value
                 return None
             
             result.sheets['Invoice']['col_amount'] = get_v('col_amount') or "N/A"
             result.sheets['Invoice']['col_qty_sf'] = get_v('col_qty_sf') or "N/A"
             
             pal_val = get_v('col_pallet_count')
             pal_regex_val = None
             for cell in row_cells:
                 if isinstance(cell.value, str) and 'pallet' in cell.value.lower():
                     pal_regex_val = extract_pallet_info(cell.value)
             
             final_p = pal_val if pal_val else pal_regex_val
             if final_p:
                  if isinstance(final_p, str):
                       num = regex_extract_number(final_p, default=0.0)
                       final_p = int(num) if num == int(num) else num
                  result.sheets['Invoice']['col_pallet_count'] = final_p


    # 2. Packing List
    packing_sheets = find_all_packing_list_sheets(wb)
    pack_data = {}
    if packing_sheets:
        matched_sheet_count += len(packing_sheets)
        result.sheet_status['PackingList'] = True
        
        for idx, p_sheet in enumerate(packing_sheets):
            p_formulas = wb_formulas[p_sheet.title]
            p_data = extract_packing_list_data(p_sheet, p_formulas, mapping_dict)
            result.packing_candidates.append({'sheet_name': p_sheet.title, 'data': p_data})
            if idx == 0: pack_data = p_data

        for k, v in pack_data.items():
            result.sheets['PackingList'][k] = v
        
        # Detect inspectable columns for packing list (use first sheet)
        pl_inspectable, pl_detection = detect_inspectable_columns(packing_sheets[0], mapping_dict)
        result.sheets['PackingList']['target_inspect_col'] = list(pl_inspectable)
        result.sheets['PackingList']['detection_info'] = pl_detection
        result.sheets['PackingList']['sheet_title'] = packing_sheets[0].title

    # 3. Contract
    contract_sheet = find_contract_sheet(wb)
    if contract_sheet:
        matched_sheet_count += 1
        result.sheet_status['Contract'] = True
        c_data = extract_contract_data(contract_sheet, wb_formulas[contract_sheet.title], mapping_dict)
        result.sheets['Contract'] = c_data
        
        # Detect inspectable columns for contract
        ct_inspectable, ct_detection = detect_inspectable_columns(contract_sheet, mapping_dict)
        result.sheets['Contract']['target_inspect_col'] = list(ct_inspectable)
        result.sheets['Contract']['detection_info'] = ct_detection
        result.sheets['Contract']['sheet_title'] = contract_sheet.title

    # Formatting & Flattening
    def format_val(v):
        if v is None or v == "N/A": return "N/A"
        if isinstance(v, (int, float)): return v
        return str(v)

    # Flatten & Track Sources
    # Invoice
    inv_name = inv_sheet.title if inv_sheet else "Unknown"
    result.col_amount = format_val(result.sheets['Invoice'].get('col_amount'))
    result.sources['col_amount'] = inv_name
    
    result.col_qty_sf = format_val(result.sheets['Invoice'].get('col_qty_sf'))
    result.sources['col_qty_sf'] = inv_name
    
    result.col_pallet_count = format_val(result.sheets['Invoice'].get('col_pallet_count'))
    result.sources['col_pallet_count'] = inv_name
    
    # Packing List (Assuming first candidate is used)
    pl = result.sheets['PackingList']
    pl_name = "Unknown"
    if packing_sheets:
         # Find the sheet that provided the data (simplified: assumes first one used for pack_data)
         pl_name = packing_sheets[0].title 
    
    result.col_qty_pcs = format_val(pl.get('col_qty_pcs'))
    result.sources['col_qty_pcs'] = pl_name
    
    result.col_net = format_val(pl.get('col_net'))
    result.sources['col_net'] = pl_name
    
    result.col_gross = format_val(pl.get('col_gross'))
    result.sources['col_gross'] = pl_name
    
    result.col_cbm = format_val(pl.get('col_cbm'))
    result.sources['col_cbm'] = pl_name

    # Pallet Strategy (Backfill)
    if result.col_pallet_count == "N/A" and pl.get('col_pallet_count'):
         result.col_pallet_count = format_val(pl['col_pallet_count'])
         result.sheets['Invoice']['col_pallet_count'] = pl['col_pallet_count'] 
         result.sources['col_pallet_count'] = pl_name # Update source to PL

    return result

def parse_filename(file_path: Path, known_ids: Set[str] = None) -> dict:
    """Parses filename to extract ID."""
    original_name = file_path.name
    extracted_id = None
    
    if known_ids:
        sorted_ids = sorted(known_ids, key=len, reverse=True)
        for k_id in sorted_ids:
            if k_id in original_name:
                extracted_id = k_id
                break
    
    if not extracted_id:
        # Regex Fallback - capture compound IDs like JLF-ISELLA26002 or simple IDs like MOTO26003E
        # Pattern: Optional(Letters-) + Letters + OptionalHyphen + Digits + OptionalTrailingLetters
        candidates = re.findall(r'((?:[A-Z]+[-_])?[A-Z]+[-_]?\d+[A-Z]*)', original_name, re.IGNORECASE)
        valid_candidates = []
        noise_prefixes = {'COPY', 'XLS', 'V', 'PART', 'REV', 'VAL', 'NUM', 'NO'}
        for c in candidates:
            upper_c = c.upper()
            prefix_match = re.match(r'^([A-Z]+)', upper_c)
            if prefix_match:
                if prefix_match.group(1) in noise_prefixes: continue
            valid_candidates.append(upper_c)
        
        if valid_candidates:
            extracted_id = max(valid_candidates, key=len)

    return {
        'original_path': file_path,
        'extracted_id': extracted_id,
        'original_name': original_name
    }

def scan_invoice_files(target_folder: Path) -> List[Dict]:
    """Scans folder for Invoice files."""
    scanned = []
    all_files = list(target_folder.glob("*.xlsx")) + list(target_folder.glob("*.xls"))
    
    for f in all_files:
        if "master" in f.name.lower(): continue
        if f.name.startswith("~$"): continue
        if f.name in ["manual_review_needed.csv", "final_invoice_data.json", "rejection_report.csv", "missing_invoices.csv", "verification_report.xlsx", "verification_report.csv"]: continue
        
        scanned.append(parse_filename(f))
    return scanned
