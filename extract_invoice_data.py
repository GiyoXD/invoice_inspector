import argparse
import re
import sys
import os
import json
import csv
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from typing import List, Dict, Set, Optional, Union
from openpyxl.styles import PatternFill

# --- Module: Helpers & Existing Logic ---

MAPPING_CONFIG_PATH = Path("mapping_config.json")

def load_mapping_config() -> dict:
    """Loads and normalizes the mapping configuration."""
    if not MAPPING_CONFIG_PATH.exists():
        print(f"Warning: {MAPPING_CONFIG_PATH} not found. Using empty mapping.")
        return {}
    
    try:
        with open(MAPPING_CONFIG_PATH, 'r', encoding='utf-8') as f:
            config = json.load(f)
            
        # Normalize mappings: Lowercase key -> Col ID
        normalized = {}
        
        # Merge source 1: header_text_mappings
        if 'header_text_mappings' in config:
            for k, v in config['header_text_mappings'].get('mappings', {}).items():
                normalized[k.lower().strip()] = v
                
        # Merge source 2: shipping_list_header_map
        if 'shipping_list_header_map' in config:
            for k, v in config['shipping_list_header_map'].get('mappings', {}).items():
                normalized[k.lower().strip()] = v
                
        return normalized
    except Exception as e:
        print(f"Error loading mapping config: {e}")
        return {}

def find_invoice_sheet(wb):
    """Finds a sheet named like 'invoice', 'inv', etc."""
    sheet_names = wb.sheetnames
    for name in sheet_names:
        lower_name = name.lower()
        if 'invoice' in lower_name or 'inv' in lower_name:
            return wb[name]
    return None

def extract_pallet_info(cell_value):
    """Extracts text to the left of 'pallet'."""
    if not isinstance(cell_value, str):
        return None
    match = re.search(r'(.*?)\s*pallet', cell_value, re.IGNORECASE)
    if match:
        return match.group(1).strip()
    return None

def identify_column_type(sheet, row_idx, col_idx, mapping_dict):
    """
    Looks upwards from (row_idx, col_idx) to find a header matching the mapping config.
    Returns the mapped col_id (e.g., 'col_qty_sf', 'col_amount') if found, else None.
    """
    # Look up to 50 rows above, or until row 1
    for r in range(row_idx - 1, 0, -1):
        cell_val = sheet.cell(row=r, column=col_idx).value
        if not cell_val:
            continue
        
        text = str(cell_val).lower().strip().replace('\n', ' ')
        
        # 1. Exact/Partial Match against Mapping Keys
        # Optimization: Check exact match first
        if text in mapping_dict:
            return mapping_dict[text]
            
        # 2. Iterate keys if needed (for partial matches if defined, but dict lookup is strict)
        # The simple lookup above covers most cases defined in the config. 
        # If the config has newlines, we might need to handle that. 
        # (Added replace \n above to help match config keys which might not have newlines or might have them differently)
        
        # Fallback Check (Legacy hardcoded if config misses common ones, or rely strictly on config?)
        # Let's rely on config + minimal fallbacks for robustness if config is broken
        
        if 'total' in text and 'value' in text: return 'col_amount'
        if 'amount' in text: return 'col_amount'
             
    return None

def find_all_packing_list_sheets(wb):
    """Finds ALL sheets that look like a packing list."""
    sheets = []
    for sheet in wb:
        title = sheet.title.lower()
        if 'pack' in title or 'weight' in title or 'detail' in title:
            sheets.append(sheet)
    return sheets

def extract_packing_list_data(sheet_values, sheet_formulas, mapping_dict) -> dict:
    """
    Extracts data from the SINGLE best 'Total' row in the Packing List sheet.
    Target fields: Pcs, Net Weight, Gross Weight, CBM.
    """
    data = {
        'col_qty_pcs': 0,
        'col_net': 0.0,
        'col_gross': 0.0,
        'col_cbm': 0.0,
        'col_pallet_count': 0
    }
    
    # 1. Find Smart Total Row
    row_idx = find_smart_total_row(sheet_values, sheet_formulas)
    
    if row_idx == -1:
        return {}
        
    # 2. Extract from that single row
    try:
        row = sheet_values[row_idx]
        for cell in row:
             if cell.value is None: continue 
             
             # Identify column type
             c_type = identify_column_type(sheet_values, row_idx, cell.column, mapping_dict)
             
             try:
                 val = cell.value
                 if val is None: continue
                 
                 # Basic float extraction
                 num = 0.0
                 val_str = str(val).strip()
                 
                 # Strategy 1: Regex for numbers
                 import re
                 m = re.search(r'(\d+(\.\d+)?)', val_str.replace(',', '').replace('$', ''))
                 if m: num = float(m.group(1))

                 # Strategy 2: Explicit Text Search for Pallets (Override/Supplement)
                 # If cell text matches "X Pallets" pattern, trust it even without column header
                 if 'pallet' in val_str.lower():
                     m_pal = re.search(r'(\d+)\s*[-_]?\s*pallet', val_str, re.IGNORECASE) or re.search(r'pallet\w*\s*[:\-]?\s*(\d+)', val_str, re.IGNORECASE) or re.search(r'(\d+)', val_str) # Fallback to any number in cell if "pallet" is present
                     if m_pal:
                         p_val = int(float(m_pal.group(1)))
                         if p_val > 0:
                             data['col_pallet_count'] = p_val
                             # If we found it via text, we don't need to treat it as a generic number
                             # But let's continue in case it's also mapped
                 
                 if num == 0.0: continue

                 if c_type == 'col_qty_pcs':
                     data['col_qty_pcs'] = int(num)
                 elif c_type == 'col_net':
                     data['col_net'] = num
                 elif c_type == 'col_gross':
                     data['col_gross'] = num
                 elif c_type == 'col_cbm':
                     data['col_cbm'] = num
                 elif c_type == 'col_pallet_count':
                     # If header says Pallet, trust the number
                     data['col_pallet_count'] = int(num)
             except:
                 pass
    except Exception as e:
        print(f"Error extracting packing list row: {e}")

    # Format and return non-zero
    res = {}
    if data['col_qty_pcs'] > 0: res['col_qty_pcs'] = data['col_qty_pcs']
    if data['col_net'] > 0: res['col_net'] = round(data['col_net'], 3)
    if data['col_gross'] > 0: res['col_gross'] = round(data['col_gross'], 3)
    if data['col_cbm'] > 0: res['col_cbm'] = round(data['col_cbm'], 4)
    if data['col_pallet_count'] > 0: res['col_pallet_count'] = data['col_pallet_count']
    
    return res

def find_contract_sheet(wb):
    """Finds a sheet named like 'contract' or 'ct'."""
    for sheet in wb:
        title = sheet.title.lower().strip()
        # 'contract' anywhere
        if 'contract' in title: 
            return sheet
        # 'ct' specific cases
        # Exact match
        if title == 'ct': return sheet
        # Boundary matches: 'ct ', 'ct-', 'ct&', 'ct_'
        if title.startswith('ct ') or title.startswith('ct-') or title.startswith('ct&') or title.startswith('ct_'):
            return sheet
        # End boundary
        if title.endswith(' ct'): return sheet
        
    return None

BLACKLIST_TERMS = ["buffalo", "cow", "leather"]

def find_smart_total_row(sheet_values, sheet_formulas) -> int:
    """
    Identifies the best 'Total' row index by checking:
    1. 'Total' label exists.
    2. Does NOT contain BLACKLIST_TERMS.
    3. Row contains formulas like '=SUM' or arithmetic '+'.
    """
    best_row_idx = -1
    max_score = 0
    
    # Iterate rows. We use sheet_values for text checking, sheet_formulas for formula checking.
    # Assuming both sheets are synced in row count.
    
    # Pre-load formulas to avoid random access issues? Iterating is safer.
    # We'll iterate sheet_formulas mainly, and peek sheet_values for labels?
    # Or iterate by index.
    
    max_row = sheet_values.max_row
    # optimization: check first 100 rows? or all? Contract sheets are small.
    # Invoice sheets might be large.
    
    # Let's iterate sheet_values to find "Total" candidates first.
    
    for row in sheet_values.iter_rows():
        row_idx = row[0].row
        
        # 1. Label Check & Blacklist Check
        has_total = False
        has_blacklist = False
        
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                v = cell.value.lower()
                # Check Blacklist first
                for term in BLACKLIST_TERMS:
                    if term in v:
                        has_blacklist = True
                        break
                if has_blacklist: break
                
                # Check Total Label
                if 'total' in v:
                    has_total = True
        
        if has_blacklist:
            continue
            
        if not has_total:
            continue
            
        # Potentially a candidate. Now check formulas in sheet_formulas at this row_idx.
        # Check Formula Criteria
        formula_score = 0
        try:
            # Get corresponding row from formula sheet
            # formula_row = sheet_formulas[row_idx] # Access by index might be slow if large?
            # It's okay for these file sizes.
            
            f_row = sheet_formulas[row_idx]
            for cell in f_row:
                if cell.value and isinstance(cell.value, str):
                    fv = cell.value.upper()
                    if str(fv).startswith('=SUM'):
                        formula_score += 2
                    elif '+' in str(fv) and '=' in str(fv): # Simple arithmetic like =C1+C2
                         formula_score += 1
        except Exception:
            pass

        # Decision Logic
        # If we have formulas, it's a strong candidate.
        # If we have NO formulas found yet, but 'Total' text is there, keep as weak candidate?
        # User requirement: "if n+ has more than 2 col using =sum, it is the row we want"
        
        current_score = 1 # Base score for having "Total"
        current_score += formula_score
        
        if current_score > max_score:
            max_score = current_score
            best_row_idx = row_idx
            
    return best_row_idx

def extract_contract_data(sheet_values, sheet_formulas) -> dict:
    """Extracts Total Quantity and Amount from Contract sheet using Smart Detection."""
    data = {'col_qty_sf': 0.0, 'col_amount': 0.0}
    
    # 1. Find Smart Total Row
    total_row_idx = find_smart_total_row(sheet_values, sheet_formulas)

    if total_row_idx == -1:
        print("  Debug: Contract Sheet - No suitable 'Total' row found.")
        return data

    # 2. Extract Values from that row (using Values sheet)
    try:
        row = sheet_values[total_row_idx]
        print(f"  Debug: Contract Total Row ({total_row_idx}): {[c.value for c in row if c.value]}")
        
        def get_float(cell):
            if cell.value:
                # Handle strings with commas
                if isinstance(cell.value, str):
                    try:
                        import re
                        m = re.search(r'(\d+(\.\d+)?)', cell.value.replace(',', '').replace('$', ''))
                        if m: return float(m.group(1))
                    except: pass
                if isinstance(cell.value, (int, float)):
                    return float(cell.value)
            return 0.0

        # Try multiple columns if 2 and 4 fail?
        # Let's inspect columns 2, 3, 4, 5, 6
        # Contract sheet usually: ... Qty ... Amount
        
        c_sqft = get_float(row[2]) # Col C (Index 2)
        c_amt = get_float(row[4])  # Col E (Index 4)
        
        # Fallback: Search row for largest numbers if specific cols fail?
        # For now just log what we see
        print(f"  Debug: Reading Col C (Idx 2): {row[2].value} -> {c_sqft}")
        print(f"  Debug: Reading Col E (Idx 4): {row[4].value} -> {c_amt}")
        
        data['col_qty_sf'] = c_sqft
        data['col_amount'] = c_amt
        
    except Exception as e:
        print(f"Error reading contract row: {e}")

    return data

def excel_data_extractor(file_path: Path) -> dict:
    """
    Extracts specific invoice data from the Excel file using Mapping Config.
    Returns a dict with a 'sheets' key containing data per sheet.
    """
    mapping_dict = load_mapping_config()
    
    # Initialize Output Structure
    result = {
        'file_path': str(file_path),
        'file_name': file_path.name,
        'invoice_id': 'Unknown',
        'sheets': {
            'Invoice': {'col_qty_sf': "N/A", 'col_amount': "N/A", 'col_pallet_count': "N/A"},
            'PackingList': {'col_qty_pcs': "N/A", 'col_net': "N/A", 'col_gross': "N/A", 'col_cbm': "N/A"},
            'Contract': {'col_qty_sf': 0.0, 'col_amount': 0.0}
        },
        # Top-level proxies for backward compatibility (GUI/Reports)
        'col_qty_sf': "N/A", 'col_amount': "N/A", 'col_qty_pcs': "N/A", 
        'col_net': "N/A", 'col_gross': "N/A", 'col_cbm': "N/A", 'col_pallet_count': "N/A",
        'verification_details': "",
        'sheet_status': {'Invoice': False, 'PackingList': False, 'Contract': False}
    }
    
    try:
        wb = load_workbook(file_path, data_only=True)
        # Load Formulas WB (Optimization: Could check if we need it first, but safer to just load)
        wb_formulas = load_workbook(file_path, data_only=False)
    except Exception as e:
        print(f"Error loading workbook: {e}")
        return result

    print(f"  Debug: Workbook Sheets: {[s.title for s in wb]}")
    
    total_sheet_count = len(wb.sheetnames)
    matched_sheet_count = 0

    # --- 1. Invoice Sheet Extraction ---
    inv_sheet = find_invoice_sheet(wb)
    
    if inv_sheet:
        matched_sheet_count += 1
        result['sheet_status']['Invoice'] = True
        # Find corresponding sheet in formulas WB
        inv_sheet_formulas = wb_formulas[inv_sheet.title]
        
        data_row_idx = find_smart_total_row(inv_sheet, inv_sheet_formulas)
                
        if data_row_idx != -1:
             found_cols = {}
             row_cells = inv_sheet[data_row_idx]
             for cell in row_cells:
                  val = cell.value
                  if val is not None:
                      c_type = identify_column_type(inv_sheet, cell.row, cell.column, mapping_dict)
                      if c_type: found_cols[c_type] = cell.column
             
             def get_v(c_type):
                 col = found_cols.get(c_type)
                 if col: return inv_sheet.cell(row=data_row_idx, column=col).value
                 return None
                 
             # Extract to Invoice Sheet dict
             result['sheets']['Invoice']['col_amount'] = get_v('col_amount') or "N/A"
             result['sheets']['Invoice']['col_qty_sf'] = get_v('col_qty_sf') or "N/A"
             
             pal_val = get_v('col_pallet_count')
             pal_regex_val = None
             for cell in row_cells:
                 if isinstance(cell.value, str) and 'pallet' in cell.value.lower():
                     pal_regex_val = extract_pallet_info(cell.value)
                     
             final_p = pal_val if pal_val else pal_regex_val
             if final_p:
                  if isinstance(final_p, str):
                       m = re.search(r'(\d+(\.\d+)?)', final_p)
                       if m: final_p = float(m.group(1))
                       if isinstance(final_p, float) and str(final_p).endswith('.0'): final_p = int(final_p)
                  result['sheets']['Invoice']['col_pallet_count'] = final_p

    # --- 2. Packing List Extraction ---
    pack_data = {}
    packing_sheets = find_all_packing_list_sheets(wb)
    result['packing_candidates'] = []
    
    if packing_sheets:
        # Mark all as identified
        matched_sheet_count += len(packing_sheets)
        result['sheet_status']['PackingList'] = True
        
        print(f"  Found {len(packing_sheets)} Packing List candidates: {[s.title for s in packing_sheets]}")
        
        # Extract from ALL candidates
        for idx, p_sheet in enumerate(packing_sheets):
            p_formulas = wb_formulas[p_sheet.title]
            p_data = extract_packing_list_data(p_sheet, p_formulas, mapping_dict)
            
            # Store candidate
            result['packing_candidates'].append({
                'sheet_name': p_sheet.title,
                'data': p_data
            })
            
            # Use First one as Default (for now)
            if idx == 0:
                pack_data = p_data
                print(f"  Defaulting to: {p_sheet.title}")

        # Populate sheets dict (Default)
        for k in ['col_qty_pcs', 'col_net', 'col_gross', 'col_cbm', 'col_pallet_count']:
            if k in pack_data:
                result['sheets']['PackingList'][k] = pack_data[k]

    # --- 3. Contract Sheet Extraction ---
    contract_sheet = find_contract_sheet(wb)
    if contract_sheet:
        matched_sheet_count += 1
        result['sheet_status']['Contract'] = True
        contract_sheet_formulas = wb_formulas[contract_sheet.title]
        c_data = extract_contract_data(contract_sheet, contract_sheet_formulas)
        result['sheets']['Contract']['col_qty_sf'] = c_data.get('col_qty_sf', 0)
        result['sheets']['Contract']['col_amount'] = c_data.get('col_amount', 0)
        print(f"  Contract Data Extracted: SQFT={c_data.get('col_qty_sf')}, Amt={c_data.get('col_amount')}")
        
    # --- Check for Unidentified Sheets ---
    if matched_sheet_count < total_sheet_count:
        msg = f"Warning: Unidentified Sheet(s) Found (Matched {matched_sheet_count} vs Total {total_sheet_count})"
        print(f"  {msg}")
        if result['verification_details']:
            result['verification_details'] += f"; {msg}"
        else:
            result['verification_details'] = msg

    # --- Formatting & Flattening for Back-Compat/Verification ---
    
    def format_val(v):
        if v is None or v == "N/A": return "N/A"
        if isinstance(v, (int, float)): return v
        return str(v)

    # 1. Formatting
    for s_name, s_data in result['sheets'].items():
        for k, v in s_data.items():
             s_data[k] = format_val(v)

    # 2. Flatten/Merge logic for Top-Level (Prefer Invoice > Packing List)
    # Invoice Data
    result['col_amount'] = result['sheets']['Invoice']['col_amount']
    result['col_qty_sf'] = result['sheets']['Invoice']['col_qty_sf']
    result['col_pallet_count'] = result['sheets']['Invoice']['col_pallet_count']
    
    # Packing List Data (Supplement Invoice if missing, or use Packing List specific)
    # Usually PCS/Weight/CBM come from Packing List.
    # But if checking generic keys, we map them here.
    
    pl = result['sheets']['PackingList']
    result['col_qty_pcs'] = pl.get('col_qty_pcs', "N/A")
    result['col_net'] = pl.get('col_net', "N/A")
    result['col_gross'] = pl.get('col_gross', "N/A")
    result['col_cbm'] = pl.get('col_cbm', "N/A")

    # Pallet Fallback (If Invoice missed it)
    if result['col_pallet_count'] == "N/A" or result['col_pallet_count'] == 0:
        if pl.get('col_pallet_count'):
             result['col_pallet_count'] = pl['col_pallet_count']
             # Backfill Invoice sheet for verification purposes
             result['sheets']['Invoice']['col_pallet_count'] = pl['col_pallet_count']
             print(f"  [Info] Using Pallet Count from Packing List: {pl['col_pallet_count']}")
    


    return result

# --- Module: Filename Parser ---

def parse_filename(file_path: Path, known_ids: Set[str] = None) -> dict:
    """
    Parses the filename to extract a Short ID.
    Priority 1: Check if any ID from 'known_ids' (Master List) exists in the filename.
    Priority 2: Use Regex to find candidates (for detecting Rejected/Unknown IDs).
    """
    original_name = file_path.name
    extracted_id = None
    
    # Priority 1: Check against Master List
    if known_ids:
        # Check for presence of known IDs in the filename
        # We need to handle potential overlaps, e.g. "INV-1" vs "INV-10". 
        # Strategy: Sort known_ids by length (descending) to match longest specific ID first.
        # Ideally this sorting should happen once outside, but for simplicity/robustness we do it here or assume caller handles.
        # Since known_ids is a Set, we must sort it to iterate.
        sorted_ids = sorted(known_ids, key=len, reverse=True)
        
        for k_id in sorted_ids:
            if k_id in original_name:
                extracted_id = k_id
                break
    
    if extracted_id:
        return {
            'original_path': file_path,
            'extracted_id': extracted_id,
            'original_name': original_name
        }

    # Priority 2: Fallback to Regex (for Rejected identification)
    # Strategy: Look for standard Invoice ID patterns.
    # Regex: Capture alphanumeric prefixes followed by optional separator and digits.
    # using re.findall to handle cases where noise like 'Part1' appears before the real ID.
    
    candidates = re.findall(r'([A-Z]+[-_]?\d+)', original_name, re.IGNORECASE)
    
    valid_candidates = []
    noise_prefixes = {'COPY', 'XLS', 'V', 'PART', 'REV', 'VAL', 'NUM', 'NO'}
    
    for c in candidates:
        upper_c = c.upper()
        # Filter: check prefix
        prefix_match = re.match(r'^([A-Z]+)', upper_c)
        if prefix_match:
            prefix = prefix_match.group(1)
            if prefix in noise_prefixes:
                continue
        
        valid_candidates.append(upper_c)
        
    if valid_candidates:
        # Prefer longest candidate (heuristic: INV-001 > V1)
        extracted_id = max(valid_candidates, key=len)

    # If simple regex fails, return None (failed_parse)
    if not extracted_id:
        pass

    return {
        'original_path': file_path,
        'extracted_id': extracted_id,
        'original_name': original_name
    }


# --- Module: Scanner ---

def scan_invoice_files(target_folder: Path) -> List[Dict]:
    """
    Scans the folder for Excel files and parses their filenames.
    Returns a list of parsed results (dictionaries).
    """
    scanned_files = []
    # Scan xlsx and xls
    all_files = list(target_folder.glob("*.xlsx")) + list(target_folder.glob("*.xls"))
    
    for f in all_files:
        # Exclude known non-invoice files
        if "master" in f.name.lower():
            continue
        if f.name in ["manual_review_needed.csv", "final_invoice_data.json", "rejection_report.csv", "missing_invoices.csv", "verification_report.xlsx", "verification_report.csv"]:
            continue
        if f.name.startswith("~$"):
            continue
            
        parsed = parse_filename(f)
        scanned_files.append(parsed)
        
    return scanned_files

# --- Module: The Auditor ---

def load_master_ids(master_path: Path) -> tuple[Set[str], Set[str]]:
    """
    Loads valid Invoice IDs from a Master List (Excel or CSV).
    Returns a tuple: (all_valid_ids, already_verified_ids)
    """
    if not master_path.exists():
        print(f"Error: Master file not found at {master_path}")
        return set(), set()
    
    try:
        if master_path.suffix.lower() == '.csv':
            df = pd.read_csv(master_path)
        else:
            df = pd.read_excel(master_path)
            
        # 1. Identify ID Column
        col_to_use = None
        cols_norm = [c.lower() for c in df.columns]
        orig_cols = list(df.columns)
        
        if 'invoice no' in cols_norm:
            col_to_use = orig_cols[cols_norm.index('invoice no')]
        elif 'invoice id' in cols_norm:
            col_to_use = orig_cols[cols_norm.index('invoice id')]
        elif 'inv id' in cols_norm:
              col_to_use = orig_cols[cols_norm.index('inv id')]
        elif 'id' in cols_norm:
               col_to_use = orig_cols[cols_norm.index('id')]
        else:
            col_to_use = orig_cols[0] # Fallback to first column

        # 2. Identify Verify State Column
        verify_col = None
        if 'verify state' in cols_norm:
            verify_col = orig_cols[cols_norm.index('verify state')]
        elif 'verified' in cols_norm:
             verify_col = orig_cols[cols_norm.index('verified')]

        # Get All IDs
        ids = set(df[col_to_use].dropna().astype(str).str.strip().unique())
        
        # Get Verified IDs
        verified_ids = set()
        if verify_col:
            # Check for True (bool) or "True" (str)
            # Normalize to string lower
            verified_mask = df[verify_col].astype(str).str.lower() == 'true'
            verified_df = df[verified_mask]
            
            if not verified_df.empty:
                verified_ids = set(verified_df[col_to_use].dropna().astype(str).str.strip().unique())

        print(f"Loaded {len(ids)} valid IDs from Master List ({master_path.name}).")
        print(f"  - Already Verified: {len(verified_ids)}")
        return ids, verified_ids
        
    except Exception as e:
        print(f"Error reading master file: {e}")
        return set(), set()

def reconcile_invoices(scanned_files: list[dict], master_ids: Set[str]) -> dict:
    """Categorize files into matched, missing, rejected, failed_parse."""
    matched = []
    rejected = []
    failed_parse = []
    
    # Track which master IDs were found
    found_master_ids = set()
    
    for file_dat in scanned_files:
        ext_id = file_dat.get('extracted_id')
        
        if not ext_id:
            failed_parse.append(file_dat)
            continue
            
        if ext_id in master_ids:
            matched.append(file_dat)
            found_master_ids.add(ext_id)
        else:
            # Found by regex but not in master -> Rejected
            rejected.append(file_dat)
            
    missing_ids = master_ids - found_master_ids
    
    return {
        'matched': matched,
        'rejected': rejected,
        'failed_parse': failed_parse,
        'missing': list(missing_ids)
    }

# --- Module: Reporting ---

REPORTS_DIR = Path("reports")

def generate_rejection_report(rejected: list[dict], failed: list[dict], missing: list[str], extraction_failed: list[dict] = None):
    """Generates separate CSV reports for rejected/failed files and missing invoices."""
    
    # Ensure reports directory exists
    REPORTS_DIR.mkdir(exist_ok=True)
    
    # Report 1: Missing Invoices
    if missing:
        missing_path = REPORTS_DIR / "missing_invoices.csv"
        try:
            with open(missing_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.DictWriter(f, fieldnames=['Missing Invoice ID'])
                writer.writeheader()
                for m_id in missing:
                    writer.writerow({'Missing Invoice ID': m_id})
            print(f"Report generated: {missing_path} ({len(missing)} entries)")
        except Exception as e:
            print(f"Error writing missing report: {e}")
    else:
        print("No missing invoices.")

    # Report 2: Rejected / Failed Files
    rejection_rows = []
    
    # Rejected (Found but wrong)
    for item in rejected:
        rejection_rows.append({
            'Original Filename': item['original_name'],
            'Extracted ID': item['extracted_id'],
            'Status': 'Unknown ID (Not in Master)'
        })
    
    # Failed (Could not parse)
    for item in failed:
        rejection_rows.append({
            'Original Filename': item['original_name'],
            'Extracted ID': 'N/A',
            'Status': 'Parse Error (Could not identify ID)'
        })

    # Extraction Failed (Found ID, but data missing/bad)
    if extraction_failed:
        for item in extraction_failed:
             rejection_rows.append({
                'Original Filename': item.get('file_name', 'Unknown'),
                'Extracted ID': item.get('invoice_id', 'Unknown'),
                'Status': 'Extraction Failed (Data Missing/N/A)'
            })
        
    if rejection_rows:
        reject_path = REPORTS_DIR / "rejection_report.csv"
        try:
            with open(reject_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.DictWriter(f, fieldnames=['Original Filename', 'Extracted ID', 'Status'])
                writer.writeheader()
                writer.writerows(rejection_rows)
            print(f"Report generated: {reject_path} ({len(rejection_rows)} entries)")
        except Exception as e:
            print(f"Error writing rejection report: {e}")
    else:
        print("No rejected or failed files. Great!")

# --- Module: Organizer ---

def organize_files(target_folder: Path, reconciliation: Dict):
    """
    Moves rejected and parsing-failed files to 'not_matched_reject'.
    """
    rejected = reconciliation['rejected']
    failed_parse = reconciliation['failed_parse']
    
    # Define Folders
    # matched_dir = target_folder / "matched_rejected"
    unmatched_dir = target_folder / "not_matched_reject"
    
    # matched_dir.mkdir(exist_ok=True)
    unmatched_dir.mkdir(exist_ok=True)
    
    def move_files(file_list, dest_dir):
        """Moves files to dest_dir and updates their 'original_path'."""
        moved_count = 0
        for item in file_list:
            src_path = item['original_path']
            dest_path = dest_dir / src_path.name
            
            try:
                # Handle overwrite if exists? For now, standard rename
                if src_path.exists() and src_path != dest_path:
                    if dest_path.exists():
                         # If destination exists, maybe append timestamp? 
                         # For now let's just overwrite or skip? 
                         # safer to overwrite in this dev env context
                         dest_path.unlink()
                         
                    src_path.rename(dest_path)
                    item['original_path'] = dest_path # Update path in memory!
                    moved_count += 1
            except Exception as e:
                print(f"  Error moving {src_path.name}: {e}")
        return moved_count

    # Move Unmatched (Rejected + Failed)
    unmatched_count = move_files(rejected + failed_parse, unmatched_dir)
    print(f"Moved {unmatched_count} files to {unmatched_dir.name}/")


# --- Module: Execution ---

def batch_rename_files(matched_files: list[dict]) -> list[dict]:
    """Renames matched files to {extracted_id}.xlsx and updates the list."""
    updated_files = []
    
    print("\nBatch Renaming Files...")
    
    for file_dat in matched_files:
        original_path = file_dat['original_path']
        ext_id = file_dat['extracted_id']
        
        # Ensure extension matches original
        suffix = original_path.suffix
        new_name = f"{ext_id}{suffix}" 
        
        new_path = original_path.parent / new_name
        
        try:
            if new_path != original_path:
                original_path.rename(new_path)
                file_dat['original_path'] = new_path # Update path in object
                file_dat['renamed'] = True
            else:
                 file_dat['renamed'] = False

            updated_files.append(file_dat)
            
        except OSError as e:
            print(f"  Error renaming {original_path.name}: {e}")
            updated_files.append(file_dat)

    print("Renaming complete.")
    return updated_files

def process_data_extraction(renamed_files: list[dict], verified_ids: Set[str] = None):
    """Calls excel_data_extractor for each file and saves results to JSON."""
    results = []
    print("\nProcessing Data Extraction...")
    
    REPORTS_DIR.mkdir(exist_ok=True) # Ensure reports dir exists
    
    if verified_ids is None:
        verified_ids = set()
        
    skipped_count = 0
    
    for file_dat in renamed_files:
        path = file_dat['original_path'] # This is the NEW path after rename
        
        # Check if already verified
        ext_id = file_dat.get('extracted_id')
        # if ext_id and ext_id in verified_ids:
            # # Skip extraction
            # skipped_count += 1
            # continue

        print(f"  Extracting: {path.name}...")
        
        data = excel_data_extractor(path)
        
        if data:
            # Enforce the ID we matched earlier (High Confidence)
            # This ensures that if extraction fails (returns 'Unknown'), we still know WHICH invoice failed.
            known_id = file_dat.get('extracted_id')
            if known_id:
                data['invoice_id'] = known_id
            
            results.append(data)

    if skipped_count > 0:
        print(f"  Skipped extraction for {skipped_count} invoices (Already Verified).")

    output_json = REPORTS_DIR / "final_invoice_data.json"
    try:
        with open(output_json, 'w', encoding='utf-8') as f:
            json.dump(results, f, indent=4)
        print(f"Final data saved to: {output_json} ({len(results)} invoices)")
    except Exception as e:
        print(f"Error saving JSON: {e}")
        
    return results

# --- Module: Verification & Reporting ---

def verify_against_master(master_path: Path, extracted_data: list[dict]):
    """
    Verifies extracted data and writes results DIRECTLY back to the Master List.
    Updates columns: VERIFY STATE, DIFF_... for all fields.
    VERIFY STATE: True if matched (all diffs within tolerance), False otherwise.
    """
    print(f"\nVerifying and Updating Master List: {master_path.name}...")
    
    if not master_path.exists():
        print(f"Error: Master list not found at {master_path}")
        return

    # 1. Load Master Data
    try:
        if master_path.suffix.lower() == '.csv':
            df_master = pd.read_csv(master_path)
        else:
            df_master = pd.read_excel(master_path)
    except Exception as e:
        print(f"Error reading master file: {e}")
        return

    # Normalize Master Columns to identify them
    col_map = {}
    for c in df_master.columns:
        cl = c.lower().strip()
        # ID
        if 'invoice' in cl or 'id' in cl:
             if 'diff' not in cl and 'verify' not in cl:
                 # Priority Logic: Prefer 'invoice no' or 'invoice id' over just 'id'
                 current_best = col_map.get('invoice_id', '')
                 # If we haven't found one yet, take it
                 if not current_best:
                     col_map['invoice_id'] = c
                 # If we have one, but this one is explicitly 'invoice no' or 'invoice id', take it
                 elif 'invoice' in cl and 'invoice' not in current_best.lower():
                      col_map['invoice_id'] = c
        # Pallet
        elif 'pallet' in cl and 'diff' not in cl:
            col_map['pallets'] = c
        # Sqft / Quantity
        elif ('sqft' in cl or 'quantity' in cl or 'qty' in cl) and 'diff' not in cl and 'pcs' not in cl:
            col_map['sqft'] = c
        # Amount
        elif ('amount' in cl or 'total' in cl or 'usd' in cl) and 'diff' not in cl:
             col_map['amount'] = c
        # PCS
        elif ('pcs' in cl or 'piece' in cl) and 'diff' not in cl:
             col_map['pcs'] = c
        # Net Weight
        elif ('net' in cl or 'nw' in cl) and 'diff' not in cl:
             col_map['net_weight'] = c
        # Gross Weight
        elif ('gross' in cl or 'gw' in cl) and 'diff' not in cl:
             col_map['gross_weight'] = c
        # CBM
        elif 'cbm' in cl and 'diff' not in cl:
             col_map['cbm'] = c

    if 'invoice_id' not in col_map:
        print("Error: Could not identify 'Invoice ID' column in Master List. Ensure column name contains 'ID' or 'Invoice'.")
        # Print detected columns for debugging
        print(f"Detected columns: {list(df_master.columns)}")
        return
        
    # Prepare Extracted Data for Merge
    extracted_map = {}
    for item in extracted_data:
        inv_id = item.get('invoice_id')
        if inv_id:
            extracted_map[inv_id] = item

    # 2. Update Data in DataFrame
    # Ensure target Diff columns exist
    diff_cols = {
        'DIFF_PALLET': 'pallets',
        'DIFF_SQFT': 'sqft',
        'DIFF_AMOUNT': 'amount',
        'DIFF_PCS': 'pcs',
        'DIFF_NET': 'net_weight',
        'DIFF_GROSS': 'gross_weight',
        'DIFF_CBM': 'cbm'
    }
    
    target_cols = ['VERIFY STATE'] + list(diff_cols.keys())
    
    for col in target_cols:
        if col not in df_master.columns:
            df_master[col] = None # Initialize if missing

    # Iterate and Update
    for index, row in df_master.iterrows():
        # Get Master Values
        m_id_key = col_map.get('invoice_id')
        m_id = str(row.get(m_id_key)).strip()
        
        extracted_item = extracted_map.get(m_id)
        
        # Optimization: Preserve existing TRUE state if no new data provided
        if not extracted_item:
            curr_state = row.get('VERIFY STATE')
            is_verified = str(curr_state).lower() == 'true'
            if is_verified:
                continue

        # Helper to clean/get numeric value
        def get_num(val, source="extracted"):
            try:
                if isinstance(val, (int, float)):
                    return float(val)
                if isinstance(val, str):
                    # Clean currency/text
                    val = val.lower().replace(',', '').replace('$', '').replace('kg', '').strip()
                    if val == 'n/a' or not val:
                        return 0.0
                    import re
                    match = re.search(r'(\d+(\.\d+)?)', val)
                    if match:
                        return float(match.group(1))
                return 0.0
            except:
                return 0.0

        diffs = {}
        all_match = True
        
        if extracted_item:
            # --- Generic Verification Logic ---
            
            # Map Sheet -> { Extracted Key : Master Internal Key }
            VERIFICATION_SCOPE = {
                'Invoice': { 'col_qty_sf': 'sqft', 'col_amount': 'amount', 'col_pallet_count': 'pallets' },
                'Contract': { 'col_qty_sf': 'sqft', 'col_amount': 'amount' },
                'PackingList': { 'col_qty_pcs': 'pcs', 'col_net': 'net_weight', 'col_gross': 'gross_weight', 'col_cbm': 'cbm', 'col_pallet_count': 'pallets' }
            }
            
            # For filling DIFF columns in Master List, we prioritize standard sources
            DIFF_SOURCE_MAP = {
                'sqft': 'Invoice', 'amount': 'Invoice', 'pallets': 'Invoice',
                'pcs': 'PackingList', 'net_weight': 'PackingList', 'gross_weight': 'PackingList', 'cbm': 'PackingList'
            }

            extracted_sheets = extracted_item.get('sheets', {})
            
            # Iterate Scope
            for sheet_name, scoped_keys in VERIFICATION_SCOPE.items():
                
                # Special Handling for PackingList Candidates (Strict Mode)
                if sheet_name == 'PackingList' and extracted_item.get('packing_candidates'):
                    candidates = extracted_item.get('packing_candidates')
                    
                    for idx, cand in enumerate(candidates):
                        c_name = cand['sheet_name']
                        c_data = cand['data']
                        
                        for e_key, m_key in scoped_keys.items():
                            # 1. Master Value
                            master_col = col_map.get(m_key)
                            m_val = 0.0
                            if master_col and master_col in row:
                                m_val = get_num(row[master_col], "master")
                            
                            # 2. Extracted Value
                            e_val = get_num(c_data.get(e_key, 0), "extracted")
                            
                            # 3. Diff
                            diff = e_val - m_val
                            
                            # 4. Check
                            if abs(diff) > 1.0:
                                all_match = False
                                readable_key = e_key.replace('_', ' ').upper()
                                msg = f"'{c_name}' {readable_key} Mismatch ({e_val} vs Master {m_val})"
                                print(f"  [Fail] {msg} for {m_id}")
                                
                                if 'verification_details' in extracted_item:
                                    if msg not in extracted_item['verification_details']:
                                        extracted_item['verification_details'] += f"; {msg}"
                                else:
                                    extracted_item['verification_details'] = msg
                                
                                # 5. Populate DIFFs (Only for Primary/First Candidate)
                                # We assume the first found candidate maps to the Master List columns
                                if idx == 0:
                                     if DIFF_SOURCE_MAP.get(m_key) == sheet_name:
                                         for d_col, d_key in diff_cols.items():
                                             if d_key == m_key:
                                                 diffs[d_col] = diff
                                                 break

                else:
                    # Standard Single Verification (Invoice, Contract, or no candidates)
                    # Skip if missing
                    if not extracted_item.get('sheet_status', {}).get(sheet_name, False):
                        continue

                    sheet_data = extracted_sheets.get(sheet_name, {})
                    
                    for e_key, m_key in scoped_keys.items():
                        # 1. Master Value
                        master_col = col_map.get(m_key)
                        m_val = 0.0
                        if master_col and master_col in row:
                            m_val = get_num(row[master_col], "master")
                        
                        # 2. Extracted Value
                        e_val = get_num(sheet_data.get(e_key, 0), "extracted")
                        
                        # 3. Diff
                        diff = e_val - m_val
                        
                        # 4. Check
                        if abs(diff) > 1.0:
                            all_match = False
                            readable_key = e_key.replace('_info', '').replace('_', ' ').upper()
                            if readable_key == 'SQFT': readable_key = 'SQFT/QTY'
                            
                            msg = f"{sheet_name} {readable_key} Mismatch ({e_val} vs Master {m_val})"
                            print(f"  [Fail] {msg} for {m_id}")
                            
                            if 'verification_details' in extracted_item:
                                if msg not in extracted_item['verification_details']:
                                    extracted_item['verification_details'] += f"; {msg}"
                            else:
                                extracted_item['verification_details'] = msg

                        # 5. Populate DIFF columns
                        if DIFF_SOURCE_MAP.get(m_key) == sheet_name:
                             for d_col, d_key in diff_cols.items():
                                 if d_key == m_key:
                                     diffs[d_col] = diff
                                     break

            verify_state = all_match
            
        else:
            verify_state = False # Not found in extraction
            
        # Update DataFrame
        df_master.at[index, 'VERIFY STATE'] = verify_state
        
        if extracted_item:
            for d_col, d_val in diffs.items():
                df_master.at[index, d_col] = d_val
        else:
            # Clear diffs if not extracted? Or leave? Safest to clear.
             for d_col in diff_cols.keys():
                df_master.at[index, d_col] = None

    # 3. Write Back to File
    try:
        if master_path.suffix.lower() == '.csv':
            df_master.to_csv(master_path, index=False)
            print(f"Updated Master List: {master_path}")
        else:
            df_master.to_excel(master_path, index=False)
            print(f"Updated Master List: {master_path}")
            
    except Exception as e:
        print(f"Error writing back to master file: {e}")

# --- Main Orchestration ---

def run_pipeline(folder_path: str, master_path: str = None):
    """
    Programmatic entry point for the pipeline.
    Returns the final list of extracted data (with verification status if master list was present).
    """
    target_folder = Path(folder_path)
    if not target_folder.exists():
         print(f"Error: Folder not found: {target_folder}")
         return []

    print(f"\nStarting Pipeline on: {target_folder.name}")
    
    # 1. Master List
    master_ids = set()
    verified_ids = set()
    master_file = None
    
    if master_path:
        master_file = Path(master_path)
        print(f"Using Master List: {master_file}")
    else:
        print("Warning: No Master List provided.")
        # Try to find one in folder?
        print(" Scanning root folder for 'Master' file (xlsx/csv)...")
        candidates = list(target_folder.glob("*Master*.xlsx")) + list(target_folder.glob("*Master*.csv"))
        if candidates:
            master_file = candidates[0]
            print(f"  Found potential master list: {master_file.name}")
        else:
            print("  No master list found in folder.")

    if master_file:
         master_ids, verified_ids = load_master_ids(master_file)
         if not master_ids:
             print("Error: Master ID list is required to proceed with reconciliation.")
             # We can continue without it if we just want to extract, but verification won't happen.
    
    # 2. Scanning
    print("\n--- Step 2: Scanning & Parsing filenames ---")
    all_files = scan_invoice_files(target_folder)
    print(f"Scanned {len(all_files)} files.")

    # 3. Reconciliation
    print("\n--- Step 3: Reconciling Invoices ---")
    reconciliation = reconcile_invoices(all_files, master_ids)
    
    matched = reconciliation['matched']
    missing = reconciliation['missing']
    rejected = reconciliation['rejected']
    failed_parse = reconciliation['failed_parse']
    
    print(f"Matched     : {len(matched)}")
    print(f"Missing     : {len(missing)}")
    print(f"Rejected    : {len(rejected)}")
    print(f"Failed Parse: {len(failed_parse)}")
    
    # 4. Reporting (Rejection / Missing)
    report_dir = target_folder / "reports"
    report_dir.mkdir(exist_ok=True)
    
    # Detailed reporting is handled by 'generate_rejection_report' at the end of the pipeline.

    # 5. Organization (Moving files)
    # print("\n--- Step 5: Organizing Files ---")
    # organize_files(target_folder, reconciliation) # Disabled per user request

    # 6. Renaming (Disabled per user request in previous session, but function exists)
    print("\n--- Step 6: Renaming Matched Files (Disabled) ---")
    # rename_matched_files(matched)

    # 7. Extraction
    print("\n--- Step 7: Extracting Data ---")
    
    final_data = [] 
    
    files_to_extract = []
    # Extract from matched files
    if matched:
        files_to_extract = matched

    # Fallback: Extract from ALL valid parsed files if no master list (otherwise matched is empty)
    if not master_ids:
        print("No Master List used -> Extracting from ALL valid parsed files.")
        files_to_extract = all_files

    if not files_to_extract:
        print("No files to process.")
    else:
        print(f"\nProcessing Data Extraction on {len(files_to_extract)} files...")
        final_data = process_data_extraction(files_to_extract, verified_ids)
        
        output_json = report_dir / "final_invoice_data.json"
        
        # Helper to convert Path to str for JSON serialization
        class PathEncoder(json.JSONEncoder):
             def default(self, obj):
                 if isinstance(obj, Path):
                     return str(obj)
                 return super().default(obj)

        try:
            with open(output_json, 'w', encoding='utf-8') as f:
                json.dump(final_data, f, indent=4, cls=PathEncoder)
            print(f"Final data saved to: {output_json.relative_to(target_folder.parent)} ({len(final_data)} invoices)")
        except Exception as e:
            print(f"Error saving JSON: {e}")

    # Quality Check (moved logs)
    print("\n--- Checking Extraction Quality ---")

    # 8. Verification
    print("\n--- Step 4 (Final): Generating Reports ---")
    # (Done above)
    print("\n--- Step 8: Verifying against Master List ---")
    if master_file and final_data:
        verify_against_master(master_file, final_data)
    else:
        print("No extracted data available for verification.")

    print("\nPipeline Complete.")
    return final_data


def main():
    parser = argparse.ArgumentParser(description="Invoice Processing Pipeline")
    parser.add_argument("--folder", help="Folder containing invoice files", default=os.getcwd())
    parser.add_argument("--master", help="Path to Master Excel List", default=None)
    
    args = parser.parse_args()
    
    run_pipeline(args.folder, args.master)

if __name__ == "__main__":
    main()
