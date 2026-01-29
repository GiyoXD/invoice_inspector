from openpyxl import load_workbook
from pathlib import Path
import sys

# Name implies it might have spaces or special chars, handle carefully
filename = "CT&INV&PL JF26001 DAF(1).xlsx"
folder = Path("process_file_dir")
path = folder / filename

if not path.exists():
    # Try finding it
    files = list(folder.glob("*.xlsx"))
    if files:
        path = files[0]
        print(f"Using found file: {path.name}")
    else:
        print("File not found.")
        sys.exit(1)

wb = load_workbook(path, data_only=True)
print(f"All Sheets: {wb.sheetnames}")

sheet = wb.active # Assuming first sheet? extraction logic uses find_invoice_sheet
# Let's try to mimic `find_invoice_sheet`
target_sheet = sheet
for sname in wb.sheetnames:
    if 'contract' in sname.lower():
        target_sheet = wb[sname]
        break

print(f"Inspecting Sheet: {target_sheet.title}")

# Find Total Row
total_found = False
total_row_idx = -1

import re
total_regex = re.compile(r'(?i)total\s*:|total\s+of\s*:|total\s+value\s*\(?usd\)?')

for row in target_sheet.iter_rows():
    for cell in row:
        if cell.value and isinstance(cell.value, str):
             if total_regex.search(cell.value) or "total" in cell.value.lower():
                 print(f"Potential Total Row {cell.row}: {cell.value}")
                 total_row_idx = cell.row
                 total_found = True
                 # Print the whole row values
                 vals = [c.value for c in row]
                 print(f"Row {cell.row} Values: {vals}")

if not total_found:
    print("No Total Row found via Regex.")
    # Print last few rows?
    print("Dumping last 5 rows:")
    max_row = target_sheet.max_row
    for r in range(max(1, max_row-5), max_row+1):
        row = target_sheet[r]
        vals = [c.value for c in row]
        print(f"Row {r}: {vals}")

# Dump rows 15-26 to find Headers and Data
print("\nDumping Rows 15-26:")
for r in range(15, 27):
    row = target_sheet[r]
    vals = [c.value for c in row]
    print(f"Row {r}: {vals}")
