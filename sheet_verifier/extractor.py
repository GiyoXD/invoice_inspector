import re
import json
from pathlib import Path
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from typing import Dict, Any, Optional, Set
from core.regex_utils import regex_extract_number, regex_extract
from core.exceptions import DataExtractionError, ErrorCode, ParsingError

# Inspectable column types (verification-critical numeric fields)
INSPECTABLE_COLS = {
    'col_qty_sf', 'col_amount', 'col_pallet_count',
    'col_qty_pcs', 'col_net', 'col_gross', 'col_cbm'
}

class SheetExtractor:
    def __init__(self):
        # Centralized Regex Configuration
        self.patterns = {
            'total_row': re.compile(r'total(\s+of)?\s*[：:]', re.IGNORECASE),
            'invoice_id': re.compile(r'(invoice\s*no|ref\s*no|inv\s*id)', re.IGNORECASE),
            'pallet_footer': re.compile(r'(\d+)\s*pallet', re.IGNORECASE)
        }
        
        # Load header mappings from config
        self.header_mappings = self._load_header_mappings()

    def extract_sheet_data(self, sheet: Worksheet, sheet_type: str = "generic") -> Dict[str, Any]:
        """
        Extracts Amount, Quantity, Pallets, and Invoice ID from a given sheet.
        Also populates target_inspect_col with detected inspectable columns.
        """
        data = {
            'row_found': -1,
            'header_row': -1,
            'target_inspect_col': set(),  # Track detected inspectable columns
            'extraction_status': 'ok',  # 'ok', 'partial', 'failed'
            'extraction_errors': [],  # List of error messages
            'values': {} # key: col_id -> val: cleaned_value
        }

        # 1. First Pass: Scan rows and collect header matches per row
        # We need 3+ header matches on the same row to confirm it as the header row
        data_row_idx = -1
        row_header_matches = {}  # row_idx -> list of (col_idx, col_id) tuples
        
        for row in sheet.iter_rows(max_row=150): 
            for cell in row:
                if not cell.value: continue
                val_str = str(cell.value).strip()
                
                # Check 1: Total Row (for Amt/Qty/Pallets)
                if data_row_idx == -1 and self.patterns['total_row'].search(val_str):
                    data_row_idx = cell.row
                
                # Check 2: Invoice ID (Header)
                if not data.get('invoice_id') and self.patterns['invoice_id'].search(val_str):
                    found_id = self._extract_id_value(sheet, cell)
                    if found_id:
                        data['invoice_id'] = found_id
                
                # Check 3: Collect header matches per row
                # We use lower() to match against normalized keys in header_mappings
                if val_str.lower() in self.header_mappings:
                    col_id = self.header_mappings[val_str.lower()]
                    if cell.row not in row_header_matches:
                        row_header_matches[cell.row] = []
                    row_header_matches[cell.row].append((cell.column, col_id))
        
        # 2. Find the header row (row with MOST header matches, minimum 3)
        header_row_idx = -1
        best_match_count = 0
        best_matches = []
        
        for row_idx, matches in row_header_matches.items():
            if len(matches) >= 3 and len(matches) > best_match_count:
                best_match_count = len(matches)
                header_row_idx = row_idx
                best_matches = matches
        
        # Populate target_inspect_col from the best header row
        for col_idx, col_id in best_matches:
            if col_id in INSPECTABLE_COLS:
                data['target_inspect_col'].add(col_id)
        
        data['header_row'] = header_row_idx
        
        # STRICT: If no header row found, mark as error
        if header_row_idx == -1:
            err_msg = f"ERROR: No header row found in sheet (need 3+ matching headers). Found rows with matches: {[(r, len(m)) for r, m in row_header_matches.items()]}"
            print(err_msg)
            data['extraction_status'] = 'failed'
            data['extraction_errors'].append('Header row not found (need 3+ matches)')
        else:
            detected = list(data['target_inspect_col'])
            print(f"  Detected inspectable columns: {detected}")

        if data_row_idx == -1:
            # CRITICAL: No Total row found - cannot extract numeric values
            err_msg = "CRITICAL: No 'Total' row found in sheet. Cannot extract Amount/Quantity values."
            print(err_msg)
            data['extraction_status'] = 'failed'
            data['extraction_errors'].append(err_msg)
            return data
        
        data['row_found'] = data_row_idx
        row_cells = sheet[data_row_idx]

        
        # 2. Extract Data from that row
        # We look for:
        # - Pallet info (text with 'pallet')
        # - Numbers/Formulas (for Amt/Qty)
        
        formula_cells = []
        
        for cell in row_cells:
            val = cell.value
            
            # Pallet Check (Text) - also detect via footer pattern
            if isinstance(val, str) and 'pallet' in val.lower():
                extracted_pal = self._extract_pallet_number(val)
                if extracted_pal is not None:
                    data['pallets'] = extracted_pal
                    # Auto-add col_pallet_count even without header
                    data['target_inspect_col'].add('col_pallet_count')
            
            # Numeric/Formula Check
            is_formula = isinstance(val, str) and val.strip().startswith('=')
            is_number = isinstance(val, (int, float))
            
            if is_formula or is_number:
                formula_cells.append(cell)

        # 3. Extract Values for ALL detected columns
        # We iterate over the best header matches to map col_idx to col_id
        # STRICT: Only extract columns we care about for verification (numeric)
        col_id_map = {c_idx: c_id for c_idx, c_id in best_matches if c_id in INSPECTABLE_COLS}
        
        # Also auto-add pallet counts if found via text patterns (handled above in extraction but need to ensure value)
        # Re-scan the row to pull values for these columns
        
        for cell in row_cells:
            if cell.column in col_id_map:
                col_id = col_id_map[cell.column]
                val = cell.value
                try:
                    clean_val = self._clean_number(val, context_col=col_id)
                    data['values'][col_id] = clean_val
                except ParsingError as e:
                    # Capture exact error
                     data['extraction_status'] = 'failed'
                     data['extraction_errors'].append(f"Value Error in {col_id}: {e.message}")
                     # Ensure we don't just omit it, maybe store as Error object or safely fail?
                     # For strict mode, we want this to propagate or be recorded. 
                     # The verifier will check if key exists. If we don't add it, it's missing.
                     # But we want to distinguish "Missing" vs "Invalid".
                     data['values'][col_id] = None # Mark as found but invalid? Or raise?
                     # Let's re-raise to be safely caught by caller? 
                     # Actually, better to store as None and let verifier see "None" -> Missing/Invalid?
                     pass

        # Handle Pallets special case (if extracted via text pattern but not header)
        if 'pallets' in data: # from text match
             data['values']['col_pallet_count'] = data['pallets']

        return data

    def _extract_pallet_number(self, text: str) -> Optional[float]:
        if not text: return None
        result = regex_extract_number(text, default=None)
        return result

    def _clean_number(self, val: Any, context_col: str = '') -> Optional[float]:
        """
        STRICT cleaning.
        - None -> None
        - Float/Int -> Float
        - String -> Attempt parse, else RAISE error.
        """
        if val is None: return None
        if isinstance(val, (int, float)): return float(val)
        
        s_val = str(val).strip()
        if not s_val: return None
        if s_val == '-' or s_val == '.': return 0.0 # Common accounting specific
        
        try:
            # Remove currency/commas
            clean = s_val.replace(',', '').replace('$', '').replace('USD', '').strip()
            return float(clean)
        except ValueError:
            # STRICT FAILURE
            raise DataExtractionError(
                error_code=ErrorCode.VALUE_PARSE_ERROR,
                message=f"Invalid number format: '{s_val}'",
                file_name="", # to be filled
                context={"column": context_col, "value": s_val}
            )

    def _find_header_type(self, sheet, row_idx, col_idx) -> Optional[str]:
        # Search upwards
        for r in range(row_idx - 1, max(0, row_idx - 50), -1):
            cell_val = sheet.cell(row=r, column=col_idx).value
            if not cell_val: continue
            
            txt = str(cell_val).lower()
            
            if 'quantity' in txt or 'qty' in txt or 'pcs' in txt or 'sqft' in txt:
                return 'quantity'
            if 'amount' in txt or 'usd' in txt or 'value' in txt:
                return 'amount'
        return None

    def _extract_id_value(self, sheet, cell) -> Optional[str]:
        """
        Extracts the ID string by searching the cell and its neighbors.
        Priority:
        1. Same Cell (stripped)
        2. Right Cell (Col + 1)
        3. Below Cell (Row + 1)
        4. Right-Below (Row + 1, Col + 1)
        """
        val = str(cell.value)
        # Remove label (Invoice No:, Ref No:, etc) using regex_extract
        clean_val = regex_extract(val, r'(?:invoice\s*no|ref\s*no|inv\s*id)\s*[:.]?\s*(.+)', group=1, default='')
        if not clean_val:
            # No match - try to use the whole value if it doesn't look like a label
            clean_val = val.strip()
            if any(x in clean_val.lower() for x in ['invoice', 'ref', 'inv']):
                clean_val = ''  # It's just a label, no ID here
        
        if len(clean_val) > 1:
            return clean_val
        
        # Define search offsets: (row_offset, col_offset)
        # We search Right, then Below, then Below-Right (optional)
        neighbors = [
            (0, 1),  # Right
            (1, 0),  # Below
            (1, 1)   # Below-Right
        ]
        
        for r_off, c_off in neighbors:
            try:
                target_cell = sheet.cell(row=cell.row + r_off, column=cell.column + c_off)
                val = target_cell.value
                if val:
                    s_val = str(val).strip()
                    if s_val:
                        return s_val
            except:
                continue
            
        return None

    def _load_header_mappings(self) -> Dict[str, str]:
        """
        Loads header text -> col_id mappings from mapping_config.json.
        Combines header_text_mappings and shipping_list_header_map.
        """
        config_path = Path("mapping_config.json")
        if not config_path.exists():
            return {}
        
        try:
            with open(config_path, 'r', encoding='utf-8') as f:
                config = json.load(f)
            
            mappings = {}
            
            # Merge source 1: header_text_mappings
            if 'header_text_mappings' in config:
                for k, v in config['header_text_mappings'].get('mappings', {}).items():
                    mappings[k.strip().lower()] = v
            
            # Merge source 2: shipping_list_header_map
            if 'shipping_list_header_map' in config:
                for k, v in config['shipping_list_header_map'].get('mappings', {}).items():
                    mappings[k.strip().lower()] = v
            
            return mappings
        except Exception as e:
            print(f"Warning: Could not load header mappings: {e}")
            return {}
