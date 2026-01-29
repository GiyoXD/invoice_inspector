import openpyxl
from pathlib import Path

def create_test_file():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Invoice"

    # Header Row (Row 5)
    headers = ["Item No", "Description", "PCS", "Quantity", "Unit Price", "Amount", "N.W (kgs)", "G.W (kgs)", "CBM", "PALLET"]
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=5, column=col_idx, value=h)

    # Data Rows
    ws.cell(row=6, column=1, value="ITEM-001")
    ws.cell(row=6, column=2, value="Test Item")
    ws.cell(row=6, column=3, value=100)
    ws.cell(row=6, column=4, value=500.5)
    ws.cell(row=6, column=5, value=10)
    ws.cell(row=6, column=6, value=5005)
    ws.cell(row=6, column=7, value=50.2) # Net
    ws.cell(row=6, column=8, value=55.5) # Gross
    ws.cell(row=6, column=9, value=1.2)  # CBM
    ws.cell(row=6, column=10, value=2)   # Pallet

    # Total Row (Row 10) - Simulating distant total row
    ws.cell(row=10, column=1, value="Total :")
    ws.cell(row=10, column=3, value=100)    # PCS
    ws.cell(row=10, column=4, value=500.5)  # Qty/Sqft
    ws.cell(row=10, column=6, value=5005)   # Amount
    ws.cell(row=10, column=7, value=50.2)   # Net
    ws.cell(row=10, column=8, value=55.5)   # Gross
    ws.cell(row=10, column=9, value=1.2)    # CBM
    ws.cell(row=10, column=10, value=2)     # Pallet
    
    # Filename must look like a valid invoice or we need a master list entry. 
    # Let's verify 'extract_invoice_data' handles just extraction if we pass it directly or via folder scanning.
    # We will use "INV-TEST-001.xlsx"
    
    path = Path("test_verification_env/INV-TEST-001.xlsx")
    path.parent.mkdir(exist_ok=True)
    wb.save(path)
    print(f"Created {path}")

if __name__ == "__main__":
    create_test_file()
